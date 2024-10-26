# excel_processor/core/excel_reader.py
from typing import Dict, Set, Optional, List, Tuple, Any
from pathlib import Path
import pandas as pd
import openpyxl
import re
from ..models.worksheet import WorksheetInfo

class ExcelReader:
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook = None
        self.formula_workbook = None
        self._sheet_cache: Dict[str, pd.DataFrame] = {}
        
    def read_workbook(self) -> Dict[str, WorksheetInfo]:
        """Read Excel workbook and extract all worksheet information."""
        try:
            if not self.file_path.exists():
                raise FileNotFoundError(f"Excel file not found: {self.file_path}")
            
            # Load workbook in both modes
            self.workbook = self._load_workbook(data_only=True)
            self.formula_workbook = self._load_workbook(data_only=False)
            
            worksheet_info = {}
            
            # Process each sheet
            for sheet_name in self.workbook.sheetnames:
                worksheet_info[sheet_name] = self._process_worksheet(sheet_name)
                self._sheet_cache[sheet_name] = worksheet_info[sheet_name].data
            
            # Process cross-sheet dependencies
            self._process_cross_sheet_dependencies(worksheet_info)
            
            return worksheet_info
            
        except Exception as e:
            raise ValueError(f"Error reading Excel file: {str(e)}")
    
    def _load_workbook(self, data_only: bool) -> openpyxl.Workbook:
        """Load Excel workbook with error handling."""
        try:
            return openpyxl.load_workbook(
                self.file_path,
                data_only=data_only,
                read_only=True
            )
        except Exception as e:
            raise ValueError(f"Error loading workbook: {str(e)}")
    
    def _process_worksheet(self, sheet_name: str) -> WorksheetInfo:
        """Process a single worksheet."""
        sheet = self.workbook[sheet_name]
        formula_sheet = self.formula_workbook[sheet_name]
        
        # Get the data range
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        # Extract headers (first row)
        headers = []
        for col in range(1, max_col + 1):
            cell_value = sheet.cell(row=1, column=col).value
            headers.append(str(cell_value) if cell_value is not None else f'Column{col}')
        
        # Extract data
        data = []
        for row in range(2, max_row + 1):
            row_data = []
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                # Handle different data types
                value = self._process_cell_value(cell)
                row_data.append(value)
            data.append(row_data)
        
        # Create DataFrame
        df = pd.DataFrame(data, columns=headers)
        
        # Extract formulas and input columns
        formulas, input_columns = self._extract_formulas_and_inputs(
            formula_sheet, headers, max_row, max_col
        )
        
        return WorksheetInfo(
            name=sheet_name,
            data=df,
            formulas=formulas,
            input_columns=input_columns
        )
    
    def _process_cell_value(self, cell: openpyxl.cell.Cell) -> Any:
        """Process cell value with type handling."""
        if cell.value is None:
            return None
            
        if cell.is_date:
            return pd.to_datetime(cell.value)
            
        if isinstance(cell.value, (int, float)):
            return cell.value
            
        # Handle other types as strings
        return str(cell.value)
    
    def _extract_formulas_and_inputs(self,
                                   sheet: openpyxl.worksheet.worksheet.Worksheet,
                                   headers: List[str],
                                   max_row: int,
                                   max_col: int) -> Tuple[Dict[str, str], Set[str]]:
        """Extract formulas and identify input columns."""
        formulas = {}
        input_columns = set()
        
        for col_idx, header in enumerate(headers, 1):
            # Check first data row for formula
            cell = sheet.cell(row=2, column=col_idx)
            if cell.formula:
                formulas[header] = f"={cell.formula}"  # Ensure formula starts with =
            else:
                input_columns.add(header)
                
            # Validate formula consistency
            if header in formulas:
                self._validate_column_formulas(
                    sheet, col_idx, formulas[header], max_row
                )
        
        return formulas, input_columns
    
    def _validate_column_formulas(self,
                                sheet: openpyxl.worksheet.worksheet.Worksheet,
                                col_idx: int,
                                first_formula: str,
                                max_row: int):
        """Validate that all formulas in a column are consistent."""
        for row in range(3, max_row + 1):  # Start from third row
            cell = sheet.cell(row=row, column=col_idx)
            if cell.formula and f"={cell.formula}" != first_formula:
                raise ValueError(
                    f"Inconsistent formulas in column {col_idx}: "
                    f"'{first_formula}' vs '={cell.formula}'"
                )
    
    def _process_cross_sheet_dependencies(self, worksheet_info: Dict[str, WorksheetInfo]):
        """Process and validate cross-sheet dependencies."""
        for sheet_name, info in worksheet_info.items():
            for col, formula in info.formulas.items():
                dependencies = self._resolve_cross_sheet_references(formula, sheet_name)
                info.dependencies[col] = dependencies
                
                # Validate dependencies
                for dep in dependencies:
                    if '!' in dep:
                        ref_sheet = dep.split('!')[0].strip("'")
                        if ref_sheet not in worksheet_info:
                            raise ValueError(
                                f"Formula in {sheet_name}.{col} references "
                                f"non-existent sheet: {ref_sheet}"
                            )
    
    def _resolve_cross_sheet_references(self, 
                                      formula: str,
                                      current_sheet: str) -> Set[str]:
        """Resolve and validate cross-sheet references."""
        references = set()
        
        # Match sheet references (Sheet1!A1 or 'Sheet Name'!A1)
        sheet_refs = re.finditer(
            r"('?[^!]+?'?)!([A-Z]+[0-9]+(?::[A-Z]+[0-9]+)?)",
            formula
        )
        
        for match in sheet_refs:
            sheet_name = match.group(1).strip("'")
            cell_ref = match.group(2)
            full_ref = f"{sheet_name}!{cell_ref}"
            
            if sheet_name not in self._sheet_cache:
                raise ValueError(f"Referenced sheet not found: {sheet_name}")
                
            references.add(full_ref)
        
        # Handle implicit references to current sheet
        local_refs = re.finditer(
            r"(?<!['A-Za-z])([A-Z]+[0-9]+(?::[A-Z]+[0-9]+)?)",
            formula
        )
        
        for match in local_refs:
            cell_ref = match.group(1)
            if '!' not in cell_ref:  # Only add if not already a full reference
                references.add(f"{current_sheet}!{cell_ref}")
        
        return references
    
    def close(self):
        """Close workbooks and clear cache."""
        if self.workbook:
            self.workbook.close()
        if self.formula_workbook:
            self.formula_workbook.close()
        self._sheet_cache.clear()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()