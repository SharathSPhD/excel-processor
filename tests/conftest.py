# tests/conftest.py
import pytest
import pandas as pd
import numpy as np
from pathlib import Path
import openpyxl
from dataclasses import dataclass, field
from typing import Dict, Set

# Create dataclass directly to avoid import cycle
@dataclass
class WorksheetInfo:
    name: str
    data: pd.DataFrame
    formulas: Dict[str, str] = field(default_factory=dict)
    input_columns: Set[str] = field(default_factory=set)
    dependencies: Dict[str, Set[str]] = field(default_factory=dict)

@pytest.fixture
def sample_config():
    return {
        'validation': {
            'enabled': True,
            'level': 'normal',
            'tolerance': 1e-10,
            'check_formulas': True,
            'check_dependencies': True
        },
        'output': {
            'format': 'csv',
            'directory': 'test_output',
            'separate_sheets': True,
            'include_metadata': True
        },
        'processing': {
            'parallel': False,
            'chunk_size': 1000,
            'max_workers': 4
        },
        'logging': {
            'level': 'DEBUG',
            'file': None
        }
    }

@pytest.fixture
def create_test_excel(tmp_path):
    def _create_excel(content: dict):
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        for sheet_name, sheet_data in content.items():
            ws = wb.create_sheet(sheet_name)
            
            # Write headers
            for col, header in enumerate(sheet_data['headers'], 1):
                ws.cell(row=1, column=col, value=header)
            
            # Write data
            for row, row_data in enumerate(sheet_data['data'], 2):
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=row, column=col, value=value)
            
            # Write formulas
            for col, formula in sheet_data.get('formulas', {}).items():
                col_idx = sheet_data['headers'].index(col) + 1
                for row in range(2, len(sheet_data['data']) + 2):
                    ws.cell(row=row, column=col_idx, value=formula)
        
        excel_path = tmp_path / "test.xlsx"
        wb.save(excel_path)
        return excel_path
    
    return _create_excel

@pytest.fixture
def simple_excel_file(create_test_excel):
    content = {
        'Sheet1': {
            'headers': ['Input', 'Formula1', 'Formula2'],
            'data': [
                [10, None, None],
                [20, None, None],
                [30, None, None]
            ],
            'formulas': {
                'Formula1': '=Input*2',
                'Formula2': '=Formula1+Input'
            }
        },
        'Sheet2': {
            'headers': ['Value', 'Reference'],
            'data': [
                [1, None],
                [2, None],
                [3, None]
            ],
            'formulas': {
                'Reference': '=Sheet1!Formula1'
            }
        }
    }
    return create_test_excel(content)

@pytest.fixture
def sample_worksheet_info():
    data = pd.DataFrame({
        'Input': [10, 20, 30],
        'Formula1': [20, 40, 60],
        'Formula2': [30, 60, 90]
    })
    
    return WorksheetInfo(
        name="Sheet1",
        data=data,
        formulas={
            'Formula1': '=Input*2',
            'Formula2': '=Formula1+Input'
        },
        input_columns={'Input'}
    )