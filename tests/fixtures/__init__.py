# tests/fixtures/__init__.py
from pathlib import Path
import pytest
import openpyxl
import pandas as pd
import yaml

FIXTURES_DIR = Path(__file__).parent

def create_complex_formulas_excel():
    """Create complex_formulas.xlsx for testing."""
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Sheet1: Basic Calculations
    ws1 = wb.create_sheet("Calculations")
    
    # Headers
    headers = ['BaseValue', 'Multiplier', 'Result1', 'Result2', 'ComplexCalc']
    for col, header in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=header)
    
    # Data
    data = [
        [100, 1.1, None, None, None],
        [200, 1.2, None, None, None],
        [300, 1.3, None, None, None]
    ]
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws1.cell(row=row_idx, column=col_idx, value=value)
    
    # Formulas
    formulas = {
        'Result1': '=BaseValue * Multiplier',
        'Result2': '=Result1 * 2',
        'ComplexCalc': '=IF(Result1>200, SUM(BaseValue:Result2), AVERAGE(BaseValue:Result2))'
    }
    
    for col_name, formula in formulas.items():
        col_idx = headers.index(col_name) + 1
        for row in range(2, len(data) + 2):
            ws1.cell(row=row, column=col_idx, value=formula)
    
    # Sheet2: Array Formulas
    ws2 = wb.create_sheet("ArrayCalcs")
    
    array_headers = ['Values', 'Condition', 'ArrayResult1', 'ArrayResult2']
    for col, header in enumerate(array_headers, 1):
        ws2.cell(row=1, column=col, value=header)
    
    array_data = [
        [10, True, None, None],
        [20, False, None, None],
        [30, True, None, None],
        [40, False, None, None]
    ]
    
    for row_idx, row_data in enumerate(array_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws2.cell(row=row_idx, column=col_idx, value=value)
    
    # Array formulas
    ws2['C2'] = '{=IF(B2:B5, A2:A5 * 2, A2:A5 / 2)}'
    ws2['D2'] = '{=SUM(IF(B2:B5, A2:A5, 0))}'
    
    excel_path = FIXTURES_DIR / 'excel_files' / 'complex_formulas.xlsx'
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_path)
    return excel_path

def create_cross_references_excel():
    """Create cross_references.xlsx for testing."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Sheet1: Data
    ws1 = wb.create_sheet("Data")
    
    data_headers = ['ID', 'Value', 'Category']
    for col, header in enumerate(data_headers, 1):
        ws1.cell(row=1, column=col, value=header)
    
    data = [
        [1, 100, 'A'],
        [2, 200, 'B'],
        [3, 300, 'A'],
        [4, 400, 'C']
    ]
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws1.cell(row=row_idx, column=col_idx, value=value)
    
    # Sheet2: Categories
    ws2 = wb.create_sheet("Categories")
    
    cat_headers = ['Category', 'Multiplier', 'Factor']
    for col, header in enumerate(cat_headers, 1):
        ws2.cell(row=1, column=col, value=header)
    
    categories = [
        ['A', 1.1, None],
        ['B', 1.2, None],
        ['C', 1.3, None]
    ]
    
    for row_idx, row_data in enumerate(categories, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws2.cell(row=row_idx, column=col_idx, value=value)
    
    # Add formula to Factor column
    for row in range(2, len(categories) + 2):
        ws2.cell(row=row, column=3, value='=Multiplier * 2')
    
    # Sheet3: Calculations
    ws3 = wb.create_sheet("Calculations")
    
    calc_headers = ['ID', 'BaseValue', 'AdjustedValue', 'FinalValue']
    for col, header in enumerate(calc_headers, 1):
        for col_idx, value in enumerate(row_data, 1):
            ws3.cell(row=row_idx, column=col_idx, value=value)
    
    # Add formulas
    calc_formulas = {
        'BaseValue': '=VLOOKUP(ID, Data!A:B, 2, FALSE)',
        'AdjustedValue': ('=BaseValue * VLOOKUP(VLOOKUP(ID, Data!A:C, 3, FALSE), '
                         'Categories!A:B, 2, FALSE)'),
        'FinalValue': ('=AdjustedValue * VLOOKUP(VLOOKUP(ID, Data!A:C, 3, FALSE), '
                      'Categories!A:C, 3, FALSE)')
    }
    
    for row_idx in range(2, len(data) + 2):
        ws3.cell(row=row_idx, column=1, value=data[row_idx-2][0])  # ID
        for col_name, formula in calc_formulas.items():
            col_idx = calc_headers.index(col_name) + 1
            ws3.cell(row=row_idx, column=col_idx, value=formula)
    
    excel_path = FIXTURES_DIR / 'excel_files' / 'cross_references.xlsx'
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_path)
    return excel_path